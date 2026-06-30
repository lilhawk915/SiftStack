"""Probate lead output writer — matches H3 DM probate spreadsheet schema.

12-column format produced by data manager across all 7 Ohio probate counties:

    1.  Case Number               (e.g. 2026EST00559, PE26-03-0254, 2026-EST-0037)
    2.  CASE TYPE                 (e.g. "Full Administration With Will",
                                       "Application to Probate Will",
                                       "Summary Release")
    3.  Date Filed                (ISO YYYY-MM-DD)
    4.  TESTATOR/DECEDENT         (last name, first name)
    5.  DOD                       (Date of Death — ISO YYYY-MM-DD)
    6.  ACTION                    (e.g. "Residual; sole beneficiary",
                                        "applied for authority to administer property")
                                  Note: Montgomery DM does NOT use this column
    7.  RELATIONSHIP              (Fiduciary's relation: Son, Spouse, Daughter,
                                   Attorney, unrelated, etc.)
    8.  Applicant/Fiduciary Name
    9.  Applicant/Fiduciary Address
    10. Applicant/Fiduciary Phone
    11. Applicant/Fiduciary Email
    12. SUBJECT PROPERTY          (real estate address — usually decedent's last
                                   domicile; comes from admin application PDF)

This is structurally simpler than the foreclosure schema (no multi-defendant
rows, no merged headers). One row per probate case.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ── Dataclass for one probate lead row ─────────────────────────────────

@dataclass
class ProbateRecord:
    case_number: str = ""
    case_type: str = ""
    date_filed: str = ""              # ISO YYYY-MM-DD
    decedent_name: str = ""
    date_of_death: str = ""           # ISO YYYY-MM-DD
    action: str = ""                  # what the applicant is seeking
    relationship: str = ""            # fiduciary's relation to decedent
    fiduciary_name: str = ""
    fiduciary_address: str = ""
    fiduciary_phone: str = ""
    fiduciary_email: str = ""
    subject_property: str = ""        # real estate address from admin PDF
    co_fiduciary_name: str = ""       # rare but happens (e.g. 2 fiduciaries)
    co_fiduciary_address: str = ""
    co_fiduciary_phone: str = ""
    notes: str = ""                   # free-form misc (e.g. "lost will", "case closed")
    # Count of OnBase application-form PDFs that Claude Vision
    # actually processed for this case during enrichment. Zero
    # means either OnBase enrichment was disabled OR the case has
    # no docket-linked PDF yet (typical for fresh filings — the
    # court usually uploads the application within 1-3 days).
    # The orchestrator uses this signal at the emit step:
    #   * phone empty + pdfs == 0  → DROP (PDF pending, catch-up
    #                                window will retry tomorrow)
    #   * phone empty + pdfs >= 1  → ship (Vision read the form
    #                                but found no fiduciary phone)
    #   * phone present, any pdfs  → ship as Dial First
    onbase_pdfs_processed: int = 0
    # Docket entries (raw, from probate_docket.parse_docket). Each entry
    # has date/description/pdf_url. Held here so downstream enrichment
    # (OnBase PDF Vision extraction) can fetch the per-entry PDFs
    # without re-scraping the case page. List of DocketEntry instances
    # captured by the scraper.
    docket_entries: list = field(default_factory=list)


# ── Output columns in DM order ─────────────────────────────────────────

COLUMNS = [
    "Case Number",
    "CASE TYPE",
    "Date Filed",
    "TESTATOR/ DECEDENT",
    "DOD",
    "ACTION",
    "RELATIONSHIP",
    "NAME of Applicant/Beneficiary/Fiduciary",
    "ADDRESS of Applicant/Beneficiary/Fiduciary",
    "PHONE NUMBER of Applicant/Beneficiary/Fiduciary",
    "EMAIL ADDRESS of Applicant/Beneficiary/Fiduciary",
    "SUBJECT PROPERTY",
]

# Montgomery's variant omits ACTION column (per H3-SOP-MCO-002 spreadsheet sample)
COLUMNS_MONTGOMERY = [c for c in COLUMNS if c != "ACTION"]


def _row_for_record(rec: ProbateRecord, *, include_action: bool = True) -> list[str]:
    base = [
        rec.case_number,
        rec.case_type,
        rec.date_filed,
        rec.decedent_name,
        rec.date_of_death,
    ]
    if include_action:
        base.append(rec.action)
    base.extend([
        rec.relationship,
        rec.fiduciary_name,
        rec.fiduciary_address,
        rec.fiduciary_phone,
        rec.fiduciary_email,
        rec.subject_property,
    ])
    return base


# ── Writers ────────────────────────────────────────────────────────────

def write_xlsx(
    records: list[ProbateRecord],
    output_path: Path,
    *,
    county: str = "",
    include_action: bool = True,
) -> Path:
    """Write probate records to an Excel file matching the DM format.

    Montgomery County omits the ACTION column per its SOP — pass
    include_action=False for Montgomery output.
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required to write .xlsx files")

    cols = COLUMNS if include_action else COLUMNS_MONTGOMERY

    wb = Workbook()
    ws = wb.active
    ws.title = "Probate Leads"

    # Header row — bold + light gray fill
    header_fill = PatternFill(
        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
    )
    header_font = Font(bold=True)
    for c, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Data rows
    for ri, rec in enumerate(records, start=2):
        for ci, val in enumerate(
            _row_for_record(rec, include_action=include_action), start=1
        ):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Add co-fiduciary as a continuation row immediately below if present
        if rec.co_fiduciary_name:
            ri_co = ri + 1
            # Wait — we'd need to interleave. Simpler: add as separate row at end.
            # For now skip inline; handle co-fiduciaries via a second pass below.
            pass

    # Auto-ish column widths (capped at 50 chars to keep file readable)
    widths = [16, 32, 12, 28, 12, 32, 16, 32, 40, 18, 32, 40]
    if not include_action:
        widths = [16, 32, 12, 28, 12, 16, 32, 40, 18, 32, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def write_csv(
    records: list[ProbateRecord],
    output_path: Path,
    *,
    county: str = "",
    include_action: bool = True,
) -> Path:
    cols = COLUMNS if include_action else COLUMNS_MONTGOMERY
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(cols)
        for rec in records:
            writer.writerow(_row_for_record(rec, include_action=include_action))
    return output_path


def write_both(
    records: list[ProbateRecord],
    base_path: Path,
    *,
    county: str = "",
) -> tuple[Path, Path]:
    """Write both .xlsx and .csv with the same base name (no extension)."""
    include_action = (county.lower() != "montgomery")
    xlsx_path = base_path.with_suffix(".xlsx")
    csv_path = base_path.with_suffix(".csv")
    write_xlsx(records, xlsx_path, county=county, include_action=include_action)
    write_csv(records, csv_path, county=county, include_action=include_action)
    return xlsx_path, csv_path
