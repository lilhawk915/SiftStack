"""H3 Homebuyers output format — matches the data manager's existing Excel layout.

Writes two artifacts from a list of CaseRecord objects:
  - {prefix}.xlsx — Excel with merged column-group headers ("Owner's Information"
    spans 5 cols, "PROPERTY LOCATION" spans 4 cols) matching the legacy
    workbook the data manager produces by hand.
  - {prefix}.csv — flat CSV with the same column order, no merged headers,
    for downstream tooling.

Multi-defendant cases produce multiple rows. Only the FIRST row of each case
carries the case-level fields (Case Number, CASE, Date Filed, Notes,
property block). Subsequent rows have only the defendant fields populated —
mirrors the human-edited format in the legacy file.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Data model ──────────────────────────────────────────────────────────

@dataclass
class Defendant:
    name: str = ""
    street: str = ""
    city: str = ""
    state: str = ""
    zip: str = ""


@dataclass
class CaseRecord:
    case_number: str
    filing_type: str = ""        # e.g. "Complaint for Foreclosure"
    date_filed: str = ""         # MM/DD/YYYY
    notes: str = ""              # free-text litigation stage / commentary
    defendants: list[Defendant] = field(default_factory=list)
    property_street: str = ""
    property_city: str = ""
    property_state: str = ""
    property_zip: str = ""
    # Optional QA flags appended to the right of the property block.
    # Help downstream filtering and DataSift upload selection.
    absentee_owner: str = ""           # "Y" if owner mailing is out-of-county
    needs_property_lookup: str = ""    # "Y" if property addr is best-guess
    # "Y" if at least one defendant is an UNKNOWN HEIRS placeholder
    # (court text like "UNKNOWN HEIRS OF X" or "JOHN DOE"). Indicates owner
    # is deceased AND no heir has stepped up — high-value lead, needs
    # skip-trace to find the actual heirs before any outreach is possible.
    heirs_unknown: str = ""
    # When heirs_unknown=Y, this is the decedent name parsed from the
    # placeholder (e.g. "MICHAEL D. JOHNSON"). Helps drive parallel
    # probate-court searches and obituary lookups.
    heirs_unknown_decedent: str = ""
    # "Y" if the property/owner address was only recovered by a deep
    # fallback (PJR PDF OCR, Complaint PDF OCR, or service-tab synthesis).
    # These cases are reachable by us but NOT by anyone else running
    # off the same court portals — so they're high-value, low-competition
    # leads. DataSift routes these to a separate, slower, personalized
    # cadence.
    deep_prospect_unreachable: str = ""
    # Which fallback recovered the address. One of:
    #   "PJR_OCR" — Warren Preliminary Judicial Report PDF + OCR
    #   "COMPLAINT_OCR" — Warren COMPLAINT PDF + OCR (tax-foreclosure
    #       branch, no PJR available)
    #   "SERVICE_TAB" — Montgomery service-tab synthesis (party-tab
    #       AJAX failed to load)
    deep_prospect_source: str = ""

    def row_count(self) -> int:
        """How many spreadsheet rows this case will occupy."""
        return max(1, len(self.defendants))


# ── Column definitions ─────────────────────────────────────────────────

# Order matches the legacy Excel exactly, with 2 QA flag columns appended.
COLUMN_HEADERS_ROW1 = [
    "Case Number",
    "CASE",
    "Date Filed",
    "Litigation Stage/ Notes",
    # Owner's Information group (5 cols, merged)
    "Owner's Information", "", "", "", "",
    # PROPERTY LOCATION group (4 cols, merged)
    "PROPERTY LOCATION", "", "", "",
    # QA flags (not merged) — 4 cols
    "Absentee?", "Property Lookup Needed?",
    "Unknown Heirs?", "Heirs Unknown - Decedent",
]

COLUMN_HEADERS_ROW2 = [
    "", "", "", "",
    "Name", "Street", "City", "State", "Zip",
    "Street", "City", "State", "Zip",
    "", "", "", "",
]

# Flat CSV header — 15 columns, no grouping
CSV_HEADER = [
    "Case Number",
    "CASE",
    "Date Filed",
    "Litigation Stage/Notes",
    "Owner Name",
    "Owner Street",
    "Owner City",
    "Owner State",
    "Owner Zip",
    "Property Street",
    "Property City",
    "Property State",
    "Property Zip",
    "Absentee Owner",
    "Property Lookup Needed",
    "Unknown Heirs",
    "Heirs Unknown - Decedent",
]

NUM_COLS = 17


# ── Row expansion ───────────────────────────────────────────────────────

def _expand(case: CaseRecord) -> list[list[str]]:
    """Expand a CaseRecord into one or more flat rows.

    First row carries case-level fields + first defendant.
    Subsequent rows have empty case-level fields, just the defendant block.
    Property block appears only on the first row.
    """
    if not case.defendants:
        case.defendants = [Defendant()]

    rows: list[list[str]] = []
    for i, d in enumerate(case.defendants):
        first = i == 0
        rows.append([
            case.case_number if first else "",
            case.filing_type if first else "",
            case.date_filed if first else "",
            case.notes if first else "",
            d.name, d.street, d.city, d.state, d.zip,
            case.property_street if first else "",
            case.property_city if first else "",
            case.property_state if first else "",
            case.property_zip if first else "",
            case.absentee_owner if first else "",
            case.needs_property_lookup if first else "",
            case.heirs_unknown if first else "",
            case.heirs_unknown_decedent if first else "",
        ])
    return rows


# ── Excel writer ────────────────────────────────────────────────────────

def write_xlsx(cases: Sequence[CaseRecord], out_path: Path) -> Path:
    """Write the merged-header Excel matching the legacy data-manager format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Foreclosures"

    # Row 1 — group headers
    for col_idx, val in enumerate(COLUMN_HEADERS_ROW1, start=1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Merge cells E1:I1 (Owner's Information) and J1:M1 (PROPERTY LOCATION).
    # The QA flag columns (N, O) stay un-merged.
    ws.merge_cells("E1:I1")
    ws.merge_cells("J1:M1")

    # Row 2 — sub-headers
    for col_idx, val in enumerate(COLUMN_HEADERS_ROW2, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    current_row = 3
    for case in cases:
        for row in _expand(case):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

    # Column widths — readable defaults
    widths = [14, 38, 12, 50, 28, 26, 16, 8, 12, 26, 16, 8, 12, 11, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze the header rows
    ws.freeze_panes = "A3"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ── CSV writer ──────────────────────────────────────────────────────────

def write_csv(cases: Sequence[CaseRecord], out_path: Path) -> Path:
    """Write a flat CSV with the 13-column header, one row per defendant."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADER)
        for case in cases:
            for row in _expand(case):
                w.writerow(row)
    return out_path


# ── Convenience ─────────────────────────────────────────────────────────

def write_both(cases: Sequence[CaseRecord], prefix: Path) -> tuple[Path, Path]:
    """Write both .xlsx and .csv at the given prefix path (no extension)."""
    xlsx = write_xlsx(cases, prefix.with_suffix(".xlsx"))
    csv_path = write_csv(cases, prefix.with_suffix(".csv"))
    return xlsx, csv_path
